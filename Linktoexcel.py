import openpyxl


def createfile(filename):
    workbook = openpyxl.Workbook()
    work = workbook.active
    work.column_dimensions['A'].width = 20
    work.column_dimensions['B'].width = 20
    work.column_dimensions['C'].width = 20
    work.column_dimensions['D'].width = 20
    workbook.save(filename=filename)
def writedata(filename, name, sec, regno, uid):
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        data = (name, sec, regno, uid)
        sheet.append(data)
        wb.save(filename)
        return "Success"
    except:
        return "Error"


def readcardno():
    try:
        wb2 = openpyxl.load_workbook("Data/student_data.xlsx")
        sheet = wb2.active
        values = [sheet.cell(row=i, column=4).value for i in range(1, sheet.max_row + 1)]
        return values
    except:
        return "Error"


def readstudname():
    try:
        wb2 = openpyxl.load_workbook("Data/student_data.xlsx")
        sheet = wb2.active
        values = [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row + 1)]
        return values
    except:
        return "Error"


def readregno():
    try:
        wb2 = openpyxl.load_workbook("Data/student_data.xlsx")
        sheet = wb2.active
        values = [sheet.cell(row=i, column=3).value for i in range(1, sheet.max_row + 1)]
        return values
    except:
        return "Error"
def readsec():
    try:
        wb2 = openpyxl.load_workbook("Data/student_data.xlsx")
        sheet = wb2.active
        values = [sheet.cell(row=i, column=2).value for i in range(1, sheet.max_row + 1)]
        return values
    except:
        return "Error"


def writecomponent(cardno, comp_ids):
    try:
        wb = openpyxl.load_workbook("Data/stock_register.xlsx")
        sheet = wb.active
        res=""
        for i in comp_ids[0:-1]:
            res=res+str(i)+","
        res=res[0:-1]
        print(res)
        data=(cardno,res)
        sheet.append(data)
        wb.save("Data/stock_register.xlsx")
        return "Success"
    except:
        return "Error"


def getcomponents(card_number):
    wb2 = openpyxl.load_workbook("Data/stock_register.xlsx")
    sheet = wb2.active
    components = [sheet.cell(row=i, column=2).value for i in range(1, sheet.max_row + 1)]
    card_data = [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row + 1)]
    result = ""


    for i in range(0, len(card_data)):
        if (card_data[i] == card_number):
            result = result + components[i] + ","
    result=result[:-1]
    result=result.split(",")
    print(type(result))
    print(result)
    return result




def deletedata(card_number):
    wb2 = openpyxl.load_workbook("Data/stock_register.xlsx")
    sheet = wb2.active
    for idxs, rows in enumerate(sheet['A']):
        if (rows.value == card_number):
            wb2.save("Data/stock_register.xlsx")
            return idxs + 1
    wb2.save("Data/stock_register.xlsx")
    return False


def delete(card_number):
    idx = deletedata(card_number)
    if (idx == False):
        return "Success"
    else:
        wb2 = openpyxl.load_workbook("Data/stock_register.xlsx")
        sheet = wb2.active
        sheet.delete_rows(idx=idx, amount=1)
        wb2.save("Data/stock_register.xlsx")
        delete(card_number)
    return "Success"


def create_stock_register():
    workbook = openpyxl.Workbook()
    work = workbook.active
    work.column_dimensions['A'].width = 20
    work.column_dimensions['B'].width = 100
    workbook.save(filename="Data/stock_register.xlsx")

